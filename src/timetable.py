import openpyxl as op
import pandas as pd
import pymorphy2
from datetime import *


'''||||||||||||||||||||||||| MAKE A DATA FRAME FOR CURRENT DAY |||||||||||||||||||||||||'''


# infinitive form for keywords
def infinitive_form(form):
    spec_chars = [",", "."]
    form = ''.join([ch for ch in form if ch not in spec_chars])
    form = form.lower()
    infinitive = morph.parse(form)[0].normal_form
    return infinitive


# correct line
def fix_line(line):
    n = len(line)
    key_words = ["лекция"]
    ignore_words = ["перерыв", "английский", "физический", "консультация"]
    for i in range(n):
        cell = line[i]
        if cell is not None:
            words_in_cell = cell.split()
            for word in words_in_cell:
                word = infinitive_form(word)
                if word in ignore_words:
                    for ind_cell in range(0, n):
                        line[ind_cell] = None
                    break
                # это много циклов - надо пофиксить (наверное)
                if word in key_words:
                    for ind_cell in range(i + 1, n):
                        line[ind_cell] = cell
                    break
    return line


# correct data
def fix_data(raw_data):
    for j in range(len(raw_data)):
        raw_data[j] = fix_line(raw_data[j])
    return raw_data


'''||||||||||||||||||||||||| MAKE A START TIME DICT  |||||||||||||||||||||||||'''


# check time format
def in_time_format(string):
    indices = [0, 1, 3, 4]

    if len(string) == 5:
        flag = True
        for i in indices:
            if not string[i].isdigit:
                flag = False
        if flag and (string[2] == "-" or string[2] == ":"):
            return True
    return False


# find time in description of class
def find_time_format(cell):
    pred_word = "$"
    key_alpha = ["с", "в", "-с"]
    if cell is not None:
        spec_chars = [",", "."]
        cell = ''.join([ch for ch in cell if ch not in spec_chars])
        words_in_cell = cell.split()
        for word in words_in_cell:
            if in_time_format(word) and pred_word in key_alpha:
                word = word[0:2] + ":" + word[3::]
                return word
            pred_word = word
        return None


class TimeTable:
    def __init__(self, week_day, sheet, week):
        # read groups
        self.groups = [sheet.cell(row=1, column=i).value for i in range(3, 9)]
        # take into account the day of the week
        parity, low_line, high_line = week[week_day]
        # make short DataFrame for day from xlsx
        self.data = [[sheet.cell(row=2 * j + parity, column=i).value for i in range(3, 9)] for j in
                     range(low_line, high_line)]
        self.data = fix_data(self.data)
        self.time = [sheet.cell(row=2 * j + parity, column=2).value for j in range(low_line, high_line)]
        # make timetable for day
        self.day = pd.DataFrame(self.data, index=self.time, columns=self.groups)
        self.day.index.name = 'Время'
        self.day.columns.name = "Группы"
        # dict time
        self.time_of_class = {}
        for group in self.groups:
            for start in self.time:
                info = self.day[group][start]
                if info is None:
                    continue
                begin = find_time_format(info)
                if begin is None:
                    begin = start[:5]
                if begin not in self.time_of_class:
                    self.time_of_class[begin] = [[group, start]]
                else:
                    self.time_of_class[begin].append([group, start])


'''||||||||||||||||||||||||| ZOOM |||||||||||||||||||||||||'''


# output number of group, ZOOM channel, ZOOM id, if it`s necessary for current time
# эта функция, вообще говоря, не нужна
def who_want_to_zoom(cur_time):
    for group in dt.groups:
        cell = dt.day[group][cur_time]
        if cell is not None:
            words_in_cell = cell.split()
            for word in words_in_cell:
                if word in zm.rooms:
                    print(cur_time, group, word, zm.channels[int(word)])
                    break


def with_zoom(message):
    if message is not None:
        words_in_message = message.split()
        for word in words_in_message:
            if word in zm.rooms:
                return message + " " + zm.channels[int(word)]
    return message


class Zoom:
    def __init__(self, sheet):
        self.channels = {}
        self.rooms = []
        for j in range(2, 12 + 1):
            room = sheet_ZOOM.cell(row=j, column=1).value
            self.rooms.append(str(int(room)))
            id_ = sheet_ZOOM.cell(row=j, column=2).value
            self.channels[room] = id_


'''||||||||||||||||||||||||| MAIN |||||||||||||||||||||||||'''


def sending_messages():
    del_keys = []
    for key in dt.time_of_class:
        ct = datetime.now().time()
        cur_time = timedelta(hours=ct.hour, minutes=ct.minute, seconds=ct.second)
        # cur_time = timedelta(hours=11, minutes=25, seconds=00)
        h, m, s = int(key[0:2]), int(key[3::]), 0
        class_time = timedelta(hours=h, minutes=m, seconds=s)
        if class_time > cur_time:
            diff = class_time - cur_time
            hh, mm, ss = [int(i) for i in str(diff).split(":")]
            diff = 60 * hh + mm
            if diff < 15:
                for group, start in dt.time_of_class[key]:
                    text_message_for(group, with_zoom(dt.day[group][start]))
                    del_keys.append(key)
    for del_key in del_keys:
        del dt.time_of_class[del_key]


def text_message_for(num_of_group, message):
    print(num_of_group, message)


if __name__ == '__main__':
    # for analysis of words
    morph = pymorphy2.MorphAnalyzer()
    while True:
        # update info of the day
        today = datetime.now()
        # open timetable (Расписание осень 2020_2021.xlsx)
        wb = op.load_workbook("Расписание осень 2020_2021.xlsx")
        sheet_main = wb['Математика + МААД']
        # set a week mods
        week_ = {0: (0, 1, 6),
                 1: (0, 6, 11),
                 2: (0, 11, 16),
                 3: (1, 16, 21),
                 4: (1, 21, 26),
                 5: (0, 26, 31),
                 6: (-1, -1, -1)}
        # input current day of week
        cur_week_day = today.weekday()
        dt = TimeTable(cur_week_day, sheet_main, week_)
        # make a dict of ZOOM id
        sheet_ZOOM = wb['ID zoom каналов']
        zm = Zoom(sheet_ZOOM)
        while today.day == datetime.now().day:
            sending_messages()

