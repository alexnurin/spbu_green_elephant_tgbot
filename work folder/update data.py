import openpyxl as op
import pandas as pd
import pymorphy2
from datetime import *


morph = pymorphy2.MorphAnalyzer()
# open timetable (Расписание осень 2020_2021.xlsx)
wb = op.load_workbook("Расписание осень 2020_2021.xlsx")

'''||||||||||||||||||||||||| ZOOM |||||||||||||||||||||||||'''
# make a dict of ZOOM id
sheet_ZOOM = wb['ID zoom каналов']
ZOOM_channels = {}
rooms = []
for j in range(2, 12 + 1):
    room = sheet_ZOOM.cell(row=j, column=1).value
    rooms.append(str(int(room)))
    id_ = sheet_ZOOM.cell(row=j, column=2).value
    ZOOM_channels[room] = id_


# output number of group, ZOOM channel, ZOOM id, if it`s necessary for current time
def who_want_to_zoom(cur_time):
    for group in groups:
        cell = day[group][cur_time]
        if cell is not None:
            words_in_cell = cell.split()
            for word in words_in_cell:
                if word in rooms:
                    print(cur_time, group, word, ZOOM_channels[int(word)])
                    break


'''||||||||||||||||||||||||| MAKE A DATA FRAME FOR CURRENT DAY |||||||||||||||||||||||||'''


# infinitive form for keywords
def infinitive_form(form):
    spec_chars = [",", "."]
    form = ''.join([ch for ch in form if ch not in spec_chars])
    form = form.lower()
    infinitive = morph.parse(form)[0].normal_form
    return infinitive


# correct line
def fix_line(line):
    n = len(line)
    key_words = ["лекция"]
    ignore_words = ["перерыв", "английский", "физическая", "консультация"]
    for i in range(n):
        cell = line[i]
        if cell is not None:
            words_in_cell = cell.split()
            for word in words_in_cell:
                word = infinitive_form(word)
                if word in ignore_words:
                    for ind_cell in range(0, n):
                        line[ind_cell] = None
                    break
                if word in key_words:
                    for ind_cell in range(i + 1, n):
                        line[ind_cell] = cell
                    break
    return line


# correct data
def fix_data(raw_data):
    for j in range(len(raw_data)):
        raw_data[j] = fix_line(raw_data[j])
    return raw_data


# open timetable
sheet = wb['Математика + МААД']
groups = [sheet.cell(row=1, column=i).value for i in range(3, 9)]
# set a week mods
week = {1: (0, 1, 6),
        2: (0, 6, 11),
        3: (0, 11, 16),
        4: (1, 16, 21),
        5: (1, 21, 26),
        6: (0, 26, 31),
        0: (-1, -1, -1)}
# input current day of week
date = datetime.now()
week_day = date.weekday()
parity, low_line, high_line = week[week_day]
# make short DataFrame for day from toy xlsx
data = [[sheet.cell(row=2*j + parity, column=i).value for i in range(3, 9)] for j in range(low_line, high_line)]
data = fix_data(data)
time = [sheet.cell(row=2*j + parity, column=2).value for j in range(low_line, high_line)]
day = pd.DataFrame(data, index=time, columns=groups)
day.index.name = 'Время'
day.columns.name = "Группы"


'''||||||||||||||||||||||||| MAKE A START TIME DICT  |||||||||||||||||||||||||'''


# check time format
def in_time_format(string):
    indices = [0, 1, 3, 4]

    if len(string) == 5:
        flag = True
        for i in indices:
            if not string[i].isdigit:
                flag = False
        if flag and (string[2] == "-" or string[2] == ":"):
            return True
    return False


# find time in description of class
def find_time_format(cell):
    pred_word = "$"
    key_alpha = ["с", "в"]
    if cell is not None:
        spec_chars = [",", "."]
        cell = ''.join([ch for ch in cell if ch not in spec_chars])
        words_in_cell = cell.split()
        for word in words_in_cell:
            if in_time_format(word) and pred_word in key_alpha:
                word = word[0:2] + ":" + word[3::]
                return word
            pred_word = word
        return None


# dict time
time_of_class = {}
for group in groups:
    for start in time:
        info = day[group][start]
        begin = find_time_format(info)
        if begin is None:
            begin = start[:5]
        if begin not in time_of_class:
            time_of_class[begin] = [[group, start]]
        else:
            time_of_class[begin].append([group, start])


'''||||||||||||||||||||||||| MAIN |||||||||||||||||||||||||'''


def main():
    for key in time_of_class:
        ct = datetime.now().time()
        cur_time = timedelta(hours=ct.hour, minutes=ct.minute, seconds=ct.second)
        h, m, s = int(key[0:2]), int(key[3::]), 0
        class_time = timedelta(hours=h, minutes=m, seconds=s)
        diff = class_time - cur_time
        print(str(diff))
        hh, mm, ss = [int(i) for i in str(diff).split(":")]
        diff = 60 * hh + mm
        if diff < 15:
            print(time_of_class[key])


main()
